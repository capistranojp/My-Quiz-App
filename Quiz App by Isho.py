import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

app = ctk.CTk()
app.title("Quiz App")
app.geometry("700x450")
app.minsize(700, 450)

quiz_file = "quizzes.xlsx"
if not os.path.exists(quiz_file):
    wb = Workbook()
    default_sheet = wb.active
    default_sheet.title = "Template"
    default_sheet['A1'] = "This is a placeholder sheet."
    wb.create_sheet("Scores")
    wb.save(quiz_file)

isho_main_frame = ctk.CTkFrame(app)
isho_main_frame.pack(fill="both", expand=True)

isho_selected_score = None


def isho_show_main(): 
    for widget in isho_main_frame.winfo_children():
        widget.destroy()

    center_frame = ctk.CTkFrame(isho_main_frame)
    center_frame.place(relx=0.5, rely=0.5, anchor="center")  # Center the frame in the window

    ctk.CTkLabel(center_frame, text="Quiz App", font=("Arial", 28)).pack(pady=20)

    button_width = 200
    button_height = 45

    ctk.CTkButton(center_frame, width=button_width, height=button_height, text="Create Quiz", command=isho_create_quiz_page).pack(pady=10)
    ctk.CTkButton(center_frame, width=button_width, height=button_height, text="Take Quiz", command=isho_select_quiz_menu).pack(pady=10)
    ctk.CTkButton(center_frame, width=button_width, height=button_height, text="Score History", command=isho_show_score_history).pack(pady=10)

def isho_select_quiz_menu():
    for widget in isho_main_frame.winfo_children():
        widget.destroy()

    center_frame = ctk.CTkFrame(isho_main_frame)
    center_frame.place(relx=0.5, rely=0.5, anchor="center")  # Center all content

    ctk.CTkLabel(center_frame, text="Select Quiz", font=("Arial", 24)).pack(pady=20)

    wb = load_workbook(quiz_file)
    quiz_names = [sheet for sheet in wb.sheetnames if sheet not in ["Template", "Scores"]]

    if not quiz_names:
        ctk.CTkLabel(center_frame, text="No quizzes available.", font=("Arial", 16)).pack(pady=10)
        ctk.CTkButton(center_frame, text="Back", width=200, height=45, command=isho_show_main).pack(pady=10)
        return

    ctk.CTkLabel(center_frame, text="Who will be taking the quiz?", font=("Arial", 16)).pack(pady=(10, 5))
    taker_entry = ctk.CTkEntry(center_frame, width=300)
    taker_entry.pack(pady=5)

    ctk.CTkLabel(center_frame, text="Select a quiz:", font=("Arial", 16)).pack(pady=(15, 5))
    quiz_var = tk.StringVar()
    quiz_dropdown = ctk.CTkOptionMenu(center_frame, variable=quiz_var, values=quiz_names, width=100)
    quiz_dropdown.pack(pady=5)
    quiz_var.set(quiz_names[0])

    def start_quiz():
        name = taker_entry.get().strip()
        quiz = quiz_var.get()
        if name and quiz:
            isho_take_quiz_page(name, quiz)

    ctk.CTkButton(center_frame, text="Start Quiz", width=200, height=45, command=start_quiz).pack(pady=15)
    ctk.CTkButton(center_frame, text="Back", width=200, height=45, command=isho_show_main).pack(pady=5)



def isho_take_quiz_page(name, quiz):
    for widget in isho_main_frame.winfo_children():
        widget.destroy()

    wb = load_workbook(quiz_file)
    sheet = wb[quiz]
    questions = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(row[:5]):
            questions.append({
                "question": row[0],
                "choices": row[1:5],
                "answer": row[5]
            })

    current_question_index = [0]
    selected_answer = tk.IntVar()
    score = [0]

    def show_question():
        for widget in isho_main_frame.winfo_children():
            widget.destroy()

        if current_question_index[0] >= len(questions):
            percentage = (score[0] / len(questions)) * 100
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            wb = load_workbook(quiz_file)
            sheet = wb["Scores"]
            sheet.append([name, quiz, f"{score[0]}/{len(questions)}", f"{percentage:.2f}%", timestamp])
            wb.save(quiz_file)

            ctk.CTkLabel(isho_main_frame, text=f"{name}, you scored {score[0]}/{len(questions)}", font=("Arial", 20)).pack(pady=20)
            ctk.CTkButton(isho_main_frame, text="Back to Menu", command=isho_show_main).pack(pady=10)
            return

        q = questions[current_question_index[0]]
        selected_answer.set(-1)

        ctk.CTkLabel(isho_main_frame, text=f"Quiz: {quiz}", font=("Arial", 16)).pack(pady=(10, 5))
        ctk.CTkLabel(isho_main_frame, text=q["question"], font=("Arial", 14), wraplength=700).pack(pady=10)

        for idx, choice in enumerate(q["choices"]):
            ctk.CTkRadioButton(isho_main_frame, text=choice, variable=selected_answer, value=idx).pack(anchor="w", padx=20)

        button_frame = ctk.CTkFrame(isho_main_frame)
        button_frame.pack(pady=20)

        def submit():
            if selected_answer.get() == q["answer"]:
                score[0] += 1
            current_question_index[0] += 1
            show_question()

        ctk.CTkButton(button_frame, text="Submit Answer", command=submit).pack(side="left", padx=10)
        ctk.CTkButton(button_frame, text="Back", command=isho_show_main).pack(side="left", padx=10)

    show_question()


# Global lists for Create Quiz
isho_quiz_name_entry = None
isho_question_entries = []
isho_choice_entries = []
isho_correct_answers = []
isho_question_canvas = None
isho_question_container = None


def isho_create_quiz_page():
    for widget in isho_main_frame.winfo_children():
        widget.destroy()

    global isho_quiz_name_entry, isho_question_entries, isho_choice_entries, isho_correct_answers
    isho_question_entries = []
    isho_choice_entries = []
    isho_correct_answers = []

    ctk.CTkLabel(isho_main_frame, text="Create Quiz", font=("Arial", 20)).pack(pady=10)
    ctk.CTkLabel(isho_main_frame, text="Enter Quiz Name:").pack()
    isho_quiz_name_entry = ctk.CTkEntry(isho_main_frame, width=300)
    isho_quiz_name_entry.pack(pady=5)

    container = ctk.CTkFrame(isho_main_frame)
    container.pack(fill="both", expand=True, pady=10)
    
    canvas = tk.Canvas(container)
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar = tk.Scrollbar(container, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")
    canvas.configure(yscrollcommand=scrollbar.set)

    isho_question_container = ctk.CTkFrame(canvas)
    canvas.create_window((0, 0), window=isho_question_container, anchor="center")

    def on_frame_configure(event):
        canvas.configure(scrollregion=canvas.bbox("all"))

    def on_mousewheel(event):
        canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    isho_question_container.bind("<Configure>", on_frame_configure)
    canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", on_mousewheel))
    canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

    def add_question():
        frame = ctk.CTkFrame(isho_question_container)
        frame.pack(pady=15)

        q_entry = ctk.CTkEntry(frame, width=600, placeholder_text="Enter question")
        q_entry.pack(pady=2)

        choices = []
        for i in range(4):
            entry = ctk.CTkEntry(frame, width=400, placeholder_text=f"Choice {chr(65+i)}")
            entry.pack(pady=2)
            choices.append(entry)

        correct_var = tk.IntVar(value=0)
        for i in range(4):
            ctk.CTkRadioButton(frame, text=f"Correct: Choice {chr(65+i)}", variable=correct_var, value=i).pack(anchor="w")

        isho_question_entries.append(q_entry)
        isho_choice_entries.append(choices)
        isho_correct_answers.append(correct_var)

        canvas.update_idletasks()
        canvas.yview_moveto(1.0)

    def remove_last_question():
        if isho_question_entries:
            isho_question_entries.pop().master.destroy()
            isho_choice_entries.pop()
            isho_correct_answers.pop()

    def save_quiz():
        name = isho_quiz_name_entry.get().strip()
        if not name:
            messagebox.showerror("Error", "Please enter a quiz name.")
            return

        wb = load_workbook(quiz_file)
        if name in wb.sheetnames:
            messagebox.showerror("Error", "Quiz already exists.")
            return

        sheet = wb.create_sheet(title=name)
        sheet.append(["Question", "ChoiceA", "ChoiceB", "ChoiceC", "ChoiceD", "CorrectIndex"])

        for q_entry, choices, correct in zip(isho_question_entries, isho_choice_entries, isho_correct_answers):
            question = q_entry.get().strip()
            choice_vals = [c.get().strip() for c in choices]
            sheet.append([question] + choice_vals + [correct.get()])

        wb.save(quiz_file)
        messagebox.showinfo("Success", "Quiz saved successfully!")
        isho_show_main()

    button_row = ctk.CTkFrame(isho_main_frame)
    button_row.pack(pady=10)

    ctk.CTkButton(button_row, text="Add Question", command=add_question).pack(side="left", padx=5)
    ctk.CTkButton(button_row, text="Remove Last Question", command=remove_last_question).pack(side="left", padx=5)
    ctk.CTkButton(button_row, text="Save Quiz", command=save_quiz).pack(side="left", padx=5)
    ctk.CTkButton(button_row, text="Back", command=isho_show_main).pack(side="left", padx=5)

    add_question()


def isho_show_score_history():
    for widget in isho_main_frame.winfo_children():
        widget.destroy()

    ctk.CTkLabel(isho_main_frame, text="Score History", font=("Arial", 20)).pack(pady=10)

    listbox = tk.Listbox(isho_main_frame, width=110, height=20)
    listbox.pack(pady=10)

    wb = load_workbook(quiz_file)
    sheet = wb["Scores"]
    scores = list(sheet.iter_rows(min_row=2, values_only=True))

    for i, row in enumerate(scores):
        entry = f"{row[4]} | {row[0]} | {row[1]} | {row[2]} | {row[3]}"
        listbox.insert(i, entry)

    def delete_selected():
        selected = listbox.curselection()
        if selected:
            index = selected[0]
            confirm = messagebox.askyesno("Delete", "Are you sure you want to delete this score?")
            if confirm:
                sheet.delete_rows(index + 2)
                wb.save(quiz_file)
                isho_show_score_history()

    def clear_all():
        confirm = messagebox.askyesno("Clear All", "Are you sure you want to clear all score history?")
        if confirm:
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for cell in row:
                    cell.value = None
            sheet.delete_rows(2, sheet.max_row)
            wb.save(quiz_file)
            isho_show_score_history()
    button_row = ctk.CTkFrame(isho_main_frame)
    button_row.pack(pady=10)
    ctk.CTkButton(button_row, text="Delete Selected Record", command=delete_selected).pack(side="left", padx=5)
    ctk.CTkButton(button_row, text="Clear All History", command=clear_all).pack(side="left", padx=5)
    ctk.CTkButton(button_row, text="Back", command=isho_show_main).pack(side="left", padx=5)


isho_show_main()
app.mainloop()
